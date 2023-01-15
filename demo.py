import os.path
import re
import time

import openpyxl
import requests
import json
from selenium import webdriver
from selenium.webdriver.edge.service import Service
from bs4 import BeautifulSoup
from lxml import etree


# 基础请求
def basicDemo():
    # 指定url: 百度
    url = "https://www.baidu.com/"
    # response发起请求, get接收响应对象
    response = requests.get(url)
    # 反馈get到的数据
    print(response.text)


# 简单网页采集器
def EasyWebPageScraper():
    # 指定url: Bing搜索
    url = "https://cn.bing.com/search"
    # Header(标头): user-agent, 可通过bing搜索F12中获取
    headers = {
        # UA(user-agent)检测: 一些网站常常通过判断 UA 来给不同的 操作系统 、不同的 浏览器 发送不同的页面，因此可能造成某些页面无法在某个浏览器中正常显示，但通过伪装 UA 可以绕过检测。 ———— 来自下列程序的爬虫搜索
        "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/108.0.0.0 Safari/537.36 Edg/108.0.1462.54"}
    # 搜索关键字
    keyword = str(input("input keyword:"))
    # 参数名字典: "q"为bing的请求参数
    params = {
        "q": keyword}
    # response发起请求, get接收响应对象, param关键参数
    response = requests.get(url=url, params=params, headers=headers)
    pageText = response.text
    fileName = keyword + ".html"
    with open(fileName, "w", encoding="UTF-8") as fp:
        fp.write(pageText)
    print(fileName, "Save Successfully")


# 百度翻译En ——> Ch
def baiduTranslate():
    url = "https://fanyi.baidu.com/sug"
    keyword = str(input("input keyword which you wanner to translate:"))
    data = {"kw": keyword}
    headers = {
        "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/108.0.0.0 Safari/537.36 Edg/108.0.1462.54"}
    # 发送请求
    response = requests.post(url, data, headers=headers)
    # 获取响应数据, json()返回obj
    dictionary = response.json()
    print(dictionary)
    fp = open("./data/translate.json", "w", encoding="UTF-8")
    # ensure-ascii (是否使用ASCII码进行编码, 中文一律False)
    json.dump(dictionary, fp, ensure_ascii=False)
    fp.close()
    print("Translate Successfully")


# 豆瓣电影
def doubanFilm():
    url = "https://movie.douban.com/j/chart/top_list"
    params = {
        "type": "24",
        "interval_id": "100:90",
        "action": "",
        "start": 0,
        "limit": 60
    }
    headers = {
        "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/108.0.0.0 Safari/537.36 Edg/108.0.1462.54"}
    response = requests.get(url, params, headers=headers)
    data = response.json()
    fp = open("./data/doubanFilm.json", "w", encoding="UTF-8")
    json.dump(data, fp, ensure_ascii=False)
    fp.close()
    print("Film Data Have Got!")


# BS4数据解析
def test():
    fp = open("data/demo.html", "r", encoding="UTF-8")
    soup = BeautifulSoup(fp, "lxml")
    # 属性定位, class属性定位
    print(soup.td)
    print(soup.find("td"))
    print(soup.find("td", class_="bgl-HP"))
    print(soup.find_all("td", class_="bgl-HP"))
    # select查找, 返回的是列表
    print(soup.select(".bgwhite > td span"))


# 三国演义爬取
def threeKingdoms():
    # 制定网址
    url = "http://guoxue.lishichunqiu.com/gdxs/sanguoyanyi/"
    # 标头UA检测
    headers = {
        "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/108.0.0.0 Safari/537.36 Edg/108.0.1462.54"
    }
    text = requests.get(url, headers=headers).text
    soup = BeautifulSoup(text, "lxml")
    ulList = soup.select(".box.no_doc > tbody > tr > td > ul")
    fp = open("./data/Three Kingdoms.docx", "w", encoding="UTF-8")
    for ul in ulList:
        title = ul.a.string
        data = ul.a["href"]
        page = requests.get(url=data, headers=headers).text
        pageSoup = BeautifulSoup(page, "lxml")
        pageSoup.find(".news_list > div > div > div")
        context = pageSoup.text
        fp.write(title + ":" + context + "\n")
        print(title, "GET SUCCESSFULLY!!!")


# xpath训练
def xpathDemo():
    tree = etree.parse("demo.html")
    # title = tree.xpath('/html/body/tr[@class="bgwhite"]')
    # //: 表示多个层级
    # title = tree.xpath("/html//tr")
    title = tree.xpath('//td[1]/text()')
    print(title)


# 58同城房源信息
def secondHandHouse():
    url = "https://bj.58.com/ershoufang/"
    headers = {
        "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/108.0.0.0 Safari/537.36 Edg/108.0.1462.54"
    }
    data = requests.get(url, headers=headers).text
    tree = etree.HTML(data)
    divList = tree.xpath("//section[@class='list']/div")
    fp = open("./data/secondHandHouse.txt", "w", encoding="UTF-8")
    for div in divList:
        # ./: 当前标签
        title = div.xpath("./a/div[2]/div/div/h3/text()")[0]
        print(title)
        fp.write(title + "\n")


# 4k美女
def fourThousandsGirls():
    url = "http://pic.netbian.com/4kmeinv/"
    headers = {
        "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/108.0.0.0 Safari/537.36 Edg/108.0.1462.54"
    }
    response = requests.get(url, headers=headers)
    # 通用处理中文乱码, encoding = "UTF-8", encoding = "iso-8859-1".decode("gbk")
    response.encoding = "GBK"
    data = response.text
    tree = etree.HTML(data)
    girlList = tree.xpath("//div[@class='slist']/ul/li")
    if not os.path.exists("./imageLibs"):
        os.mkdir("./imageLibs")
    for girl in girlList:
        imageURL = "http://pic.netbian.com" + girl.xpath("./a/img/@src")[0]
        imageName = girl.xpath("./a/img/@alt")[0] + ".jpg"
        imageData = requests.get(imageURL, headers=headers).content
        imagePath = "imageLibs/" + imageName
        with open(imagePath, "wb") as fp:
            fp.write(imageData)
        print(imageName, "已下载!")


# 全国城市名称1
def cityNameOne():
    url = "http://www.aqistudy.cn/historydata/"
    headers = {
        "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/108.0.0.0 Safari/537.36 Edg/108.0.1462.54"
    }
    response = requests.get(url, headers=headers).text
    tree = etree.HTML(response)
    cityList = tree.xpath("//div[@class='bottom']/ul/li")
    fp = open("./data/city.text", "w", encoding="UTF-8")
    for city in cityList:
        name = city.xpath("./a/text()")[0]
        print(name, "Success!")
        fp.write(name + "\n")
    cityList = tree.xpath("//div[@class='bottom']/ul/div[2]/li")
    for city in cityList:
        name = city.xpath("./a/text()")[0]
        print(name, "Success!")
        fp.write(name + "\n")
    fp.close()
    print("All data has been got~")


# 全国城市名称2
def cityNameTwo():
    url = "http://www.aqistudy.cn/historydata/"
    headers = {
        "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/108.0.0.0 Safari/537.36 Edg/108.0.1462.54"
    }
    response = requests.get(url, headers=headers).text
    tree = etree.HTML(response)
    cityList = tree.xpath("//div[@class='bottom']/ul/li | //div[@class='bottom']/ul/div[2]/li")
    fp = open("./data/city.text", "w", encoding="UTF-8")
    for city in cityList:
        name = city.xpath("./a/text()")[0]
        print(name, "Success!")
        fp.write(name + "\n")
    fp.close()


# 代理服务器
def broker():
    url = "https://uutool.cn/local-ip/"
    headers = {
        "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/108.0.0.0 Safari/537.36 Edg/108.0.1462.54"
    }
    response = requests.get(url, headers=headers, proxies={"https": "222.122.202.173:8080"}).text
    with open("data/ip.html", "w", encoding="UTF-8") as fp:
        fp.write(response)
    print("Successfully!")


# 懂车帝数据爬取
def car():
    # 实例化浏览器对象
    browser = webdriver.Edge(executable_path="edgeDriver.exe")
    # 让浏览器发起指定的url请求
    browser.get("https://www.dongchedi.com/sales/sale-energy-202207-x-x-x-x")
    browser.execute_script("window.scrollTo(0, 100000)")
    time.sleep(18)
    # 获取当前页面源码
    response = browser.page_source
    path = r"C:\Users\29646\Desktop"
    os.chdir(path)
    workbook = openpyxl.load_workbook("懂车帝新能源汽车7月销.xlsx")
    sheet = workbook.active
    tree = etree.HTML(response)
    carList = tree.xpath("//ol[@class='tw-mt-12']/li")
    data = []
    for i in carList:
        list = ["7月"]
        name = i.xpath("./div[3]/div[1]/a/text()")[0]
        list.append(name)
        type = i.xpath("./div[3]/div[1]/span/text()")[0]
        list.append(type)
        price = i.xpath("./div[3]/p/text()")[0]
        list.append(price)
        sales = i.xpath("./div[4]/div/p/text()")[0]
        list.append(sales)
        data.append(list)
    for i in data:
        sheet.append(i)
        print(i)
    workbook.save("懂车帝新能源汽车7月销.xlsx")
    browser.quit()


# 懂车帝数据爬取(可用)
def dongchedi():
    # 实例化浏览器对象, edgeDriver.exe为浏览器的启动器, 实际版本需要与自己的浏览器版本相对应
    # 参考网址: https://blog.csdn.net/kenny_pj/article/details/103646745
    browser = webdriver.Edge(executable_path="edgeDriver.exe")
    # 让浏览器发起指定的url请求
    # url如下
    url = ["https://www.dongchedi.com/sales/sale-energy-202207-x-x-x-x",
           "https://www.dongchedi.com/sales/sale-energy-202208-x-x-x-x",
           "https://www.dongchedi.com/sales/sale-energy-202209-x-x-x-x",
           "https://www.dongchedi.com/sales/sale-energy-202210-x-x-x-x",
           "https://www.dongchedi.com/sales/sale-energy-202211-x-x-x-x",
           "https://www.dongchedi.com/sales/sale-energy-202212-x-x-x-x",
           "https://www.dongchedi.com/sales/sale-x-202207-x-x-x-x",
           "https://www.dongchedi.com/sales/sale-x-202208-x-x-x-x",
           "https://www.dongchedi.com/sales/sale-x-202209-x-x-x-x",
           "https://www.dongchedi.com/sales/sale-x-202210-x-x-x-x",
           "https://www.dongchedi.com/sales/sale-x-202211-x-x-x-x",
           "https://www.dongchedi.com/sales/sale-x-202212-x-x-x-x"
           ]
    # 根据上方内容进行手动切换
    browser.get("https://www.dongchedi.com/sales/sale-energy-202207-x-x-x-x")
    browser.execute_script("window.scrollTo(0, 100000)")
    # 程序挂起18秒, 期间需要手动下翻网页
    time.sleep(18)
    # 获取当前页面源码
    response = browser.page_source
    # 更改为自己的桌面路径
    path = r"C:\Users\29646\Desktop"
    os.chdir(path)
    # 优先在桌面创建文件夹:懂车帝新能源汽车7月销.xlsx,懂车帝新能源汽车8月销.xlsx,懂车帝新能源汽车9月销.xlsx,懂车帝新能源汽车10月销.xlsx,懂车帝新能源汽车11月销.xlsx,懂车帝新能源汽车12月销.xlsx,
    # 懂车帝汽车7月销.xlsx,懂车帝汽车8月销.xlsx,懂车帝汽车9月销.xlsx,懂车帝汽车10月销.xlsx,懂车帝汽车11月销.xlsx,懂车帝汽车12月销.xlsx
    workbook = openpyxl.load_workbook("懂车帝新能源汽车7月销.xlsx")
    sheet = workbook.active
    tree = etree.HTML(response)
    carList = tree.xpath("//ol[@class='tw-mt-12']/li")
    data = []
    for i in carList:
        # 月份根据文件夹对应月份进行更改
        list = ["7月"]
        name = i.xpath("./div[3]/div[1]/a/text()")[0]
        list.append(name)
        type = i.xpath("./div[3]/div[1]/span/text()")[0]
        list.append(type)
        price = i.xpath("./div[3]/p/text()")[0]
        list.append(price)
        sales = i.xpath("./div[4]/div/p/text()")[0]
        list.append(sales)
        data.append(list)
    for i in data:
        sheet.append(i)
        print(i)
    # 与上述的workbook = openpyxl.load_workbook("懂车帝新能源汽车7月销.xlsx")文件夹内文件名称对应
    workbook.save("懂车帝新能源汽车7月销.xlsx")
    browser.quit()


def driver():
    # 实例化浏览器对象
    browser = webdriver.Edge(executable_path="./edgeDriver.exe")
    # 让浏览器发起指定的url请求
    browser.get("https://www.dongchedi.com/sales/sale-x-202207-x-x-x-x")
    response = browser.page_source
    workbook = openpyxl.load_workbook("招式表.xlsx")
    sheet = workbook.active
    tree = etree.HTML(response)
    carList = tree.xpath("//ol[@class='tw-mt-12']/li")


# 测试时间
def timeme():
    start = time.time()
    time.sleep(5)
    end = time.time()
    print(start)
    print("time:", end - start)


# 正则表达式数据匹配
# 宝可梦种族值数据爬取(个人数据分析作业使用)
def pokemonValue():
    path = r"E:\Study\PyCharm\PycharmProjects\Training\Self-Learning\Data Analysis"
    os.chdir(path)
    workbook = openpyxl.load_workbook("图鉴.xlsx")
    sheet = workbook.active
    url = "https://wiki.52poke.com/wiki/%E7%A7%8D%E6%97%8F%E5%80%BC%E5%88%97%E8%A1%A8%EF%BC%88%E7%AC%AC%E4%B9%9D%E4%B8%96%E4%BB%A3%EF%BC%89"
    headers = {
        "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/108.0.0.0 Safari/537.36 Edg/108.0.1462.54"
    }
    text = requests.get(url, headers=headers).text
    ex = '<tr class="bgwhite">.*?title="(.*?)".*?<td class="bgl-HP">(.*?)\n.*?<td class="bgl-攻击">(.*?)\n.*?<td class="bgl-防御">(.*?)\n.*?<td class="bgl-特攻">(.*?)\n.*?<td class="bgl-特防">(.*?)\n.*?<td class="bgl-速度">(.*?)\n.*?</tr>'
    dataPokemon = re.findall(ex, text, re.S)
    # 需求: 还需要将文本内容写入Excel中, 现在只知道如何写入txt文本
    data = []
    for pokemon in dataPokemon:
        list = []
        for i in pokemon:
            list.append(i)
        print(list)
        data.append(list)
    for i in data:
        sheet.append(i)
        print(i)
    workbook.save("图鉴.xlsx")


# 宝可梦招式表
def pokemonMove():
    # 实例化浏览器
    browser = webdriver.Edge(executable_path="edgeDriver.exe")
    # 获取网页信息
    browser.get("https://wiki.52poke.com/wiki/%E6%8B%9B%E5%BC%8F%E5%88%97%E8%A1%A8")
    # 缓冲1s
    time.sleep(1)
    response = browser.page_source
    # 读取路径
    path = r"E:\Study\PyCharm\PycharmProjects\Training\Self-Learning\Data Analysis"
    os.chdir(path)
    # 寻找文件
    workbook = openpyxl.load_workbook("招式表.xlsx")
    sheet = workbook.active
    tree = etree.HTML(response)
    moveList = tree.xpath(
        "//*[@id='mw-content-text']/div[1]/table[2]/tbody/tr | //*[@id='mw-content-text']/div[1]/table[3]/tbody/tr | //*[@id='mw-content-text']/div[1]/table[4]/tbody/tr | //*[@id='mw-content-text']/div[1]/table[5]/tbody/tr | //*[@id='mw-content-text']/div[1]/table[6]/tbody/tr | //*[@id='mw-content-text']/div[1]/table[7]/tbody/tr | //*[@id='mw-content-text']/div[1]/table[8]/tbody/tr | //*[@id='mw-content-text']/div[1]/table[9]/tbody/tr | //*[@id='mw-content-text']/div[1]/table[10]/tbody/tr")
    data = []
    for move in moveList:
        list = []
        name = move.xpath("normalize-space(./td[2]/a/text())")
        list.append(name)
        attribute = move.xpath("normalize-space(./td[5]/text())")
        list.append(attribute)
        type = move.xpath("normalize-space(./td[6]/a/text())")
        list.append(type)
        power = move.xpath("normalize-space(./td[7]/text())")
        list.append(power)
        hit = move.xpath("normalize-space(./td[8]/text())")
        list.append(hit)
        PP = move.xpath("normalize-space(./td[9]/text())")
        list.append(PP)
        design = move.xpath("normalize-space(./td[10]/text())")
        list.append(design)
        print(list)
        data.append(list)
    for i in data:
        sheet.append(i)
    workbook.save("招式表.xlsx")
    browser.quit()


# 宝可梦属性表
def pokemonAttribute():
    # 实例化浏览器
    browser = webdriver.Edge(executable_path="edgeDriver.exe")
    # 需要爬取的网页的网址
    url = "https://wiki.52poke.com/wiki/%E5%A6%99%E8%9B%99%E7%A7%8D%E5%AD%90"
    # 读取需要存放的文件的路径
    path = r"E:\Study\PyCharm\PycharmProjects\Training\Self-Learning\Data Analysis"
    os.chdir(path)
    # 初始化需要存放数据的文件
    workbook = openpyxl.load_workbook("属性表.xlsx")
    # 初始化sheet
    sheet = workbook.active
    while True:
        # 获取网页信息
        browser.get(url)
        # 获取网页HTML信息
        response = browser.page_source
        # 实例化etree获取HTML信息
        tree = etree.HTML(response)
        # 爬取的数据的xpath路径, 并生成为List列表类型
        List = tree.xpath("//*[@id='mw-content-text']/div[1]/table[2]/tbody | //*[@id='mw-content-text']/div[1]/table[2]/tbody/tr[1]/td/table/tbody")
        # data暂时存放数据列表, 方便后面存到文件中
        for i in List:
            # list 用于存放当前所爬取到的数据
            list = []
            # 细化需要爬取的数据xpath
            name = i.xpath("normalize-space(./tr[1]/td/table/tbody/tr/td[1]/span/b)")
            # 将上方爬取到的数据存放至list中
            list.append(name)
            # 细化需要爬取的数据xpath
            attributeOne = i.xpath("normalize-space(./tr[3]/td[1]/table/tbody/tr/td/table/tbody/tr/td/a[1]/span/span[2])")
            # 将上方爬取到的数据存放至list中
            list.append(attributeOne)
            # 细化需要爬取的数据xpath
            attributeTwo = i.xpath("normalize-space(./tr[3]/td[1]/table/tbody/tr/td/table/tbody/tr/td/a[2]/span/span[2])")
            # 将上方爬取到的数据存放至list中
            list.append(attributeTwo)
            print(list)
            # 将完整的list列表存放至data数据中
            sheet.append(list)
            workbook.save("属性表.xlsx")
        url = "https://wiki.52poke.com/" + str(tree.xpath("//*[@id='mw-content-text']/div[1]/table[1]/tbody/tr[2]/td[2]/table/tbody/tr/td[3]/a/@href")[0])

    # 实例化浏览器关闭
    browser.quit()


# 宝可梦道具列表
def pokemonProp():
    # 需要爬取的网页的网址
    url = ""
    # 实例化浏览器
    browser = webdriver.Edge(executable_path="edgeDriver.exe")
    # 获取网页信息
    browser.get(url)
    # 获取网页HTML信息
    response = browser.page_source
    # 读取需要存放的文件的路径
    path = r"E:\Study\PyCharm\PycharmProjects\Training\Self-Learning\Data Analysis\data"
    os.chdir(path)
    # 初始化需要存放数据的文件
    workbook = openpyxl.load_workbook("道具表.xlsx")
    # 初始化sheet
    sheet = workbook.active
    # 实例化etree获取HTML信息
    tree = etree.HTML(response)
    # 爬取的数据的xpath路径, 并生成为List列表类型
    List = tree.xpath("")
    # data暂时存放数据列表, 方便后面存到文件中
    data = []
    for i in List:
        # list 用于存放当前所爬取到的数据
        list = []
        # 细化需要爬取的数据xpath
        x = i.xpath("")
        # 将上方爬取到的数据存放至list中
        list.append(x)
        # 将完整的list列表存放至data数据中
        data.append(list)
    for i in data:
        sheet.append(i)
    # 保存数据
    workbook.save("道具表.xlsx")
    # 实例化浏览器关闭
    browser.quit()


def base():
    # 需要爬取的网页的网址
    url = ""
    # 实例化浏览器
    browser = webdriver.Edge(executable_path="edgeDriver.exe")
    # 获取网页信息
    browser.get(url)
    # 获取网页HTML信息
    response = browser.page_source
    # 读取需要存放的文件的路径
    path = r"E:\Study\PyCharm\PycharmProjects\Training\Self-Learning\Data Analysis"
    os.chdir(path)
    # 初始化需要存放数据的文件
    workbook = openpyxl.load_workbook("招式表.xlsx")
    # 初始化sheet
    sheet = workbook.active
    # 实例化etree获取HTML信息
    tree = etree.HTML(response)
    # 爬取的数据的xpath路径, 并生成为List列表类型
    List = tree.xpath("")
    # data暂时存放数据列表, 方便后面存到文件中
    data = []
    for i in List:
        # list 用于存放当前所爬取到的数据
        list = []
        # 细化需要爬取的数据xpath
        x = i.xpath("")
        # 将上方爬取到的数据存放至list中
        list.append(x)
        # 将完整的list列表存放至data数据中
        data.append(list)
    for i in data:
        sheet.append(i)
    # 保存数据
    workbook.save("")
    # 实例化浏览器关闭
    browser.quit()


if __name__ == '__main__':
    pokemonAttribute()
